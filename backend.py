import os

import openpyxl
from config import *
from datetime import datetime

xl_size_pos = {}
for size, letter in zip(sizes[1:], size_letters):
    xl_size_pos[size] = letter

FOLDER = os.path.join(os.getcwd(), str(datetime.today().year))
if not os.path.exists(FOLDER):
    os.makedirs(FOLDER)
SHEET_EXCEPTIONS = ['Шаблон', 'Final']


def get_path(name):
    return os.path.join(FOLDER, f"{name}.xlsx")


def copy_sheet_from_template(wb, template_name, sheet_name):
    source = wb[template_name]
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        ws = wb.copy_worksheet(source)
        ws.title = sheet_name
        return ws


def load_workbook(name: str):
    if not os.path.exists(get_path(name)):
        wb = openpyxl.load_workbook('template_full.xlsx')
        wb.save(get_path(name))
        wb.close()
    return openpyxl.load_workbook(get_path(name))


def save_workbook(wb, name: str):
    wb.save(get_path(name))
    wb.close()


def delete_workbook(name: str):
    if os.path.exists(get_path(name)):
        os.remove(get_path(name))


def read_workbook(name: str):
    if os.path.exists(get_path(name)):
        with open(get_path(name), "rb") as template_file:
            return template_file.read()


def fill_xl_student(result: dict):
    wb = load_workbook(result['second_name'])
    ws = copy_sheet_from_template(wb, 'Шаблон', result['second_name'] + ' ' + result['first_name'])
    for i, position in enumerate(result['positions'], start=4):
        ws[f"A{i}"] = position['title']
        ws[f"B{i}"] = position['price']

        if position['size'] == NOT_CHOSEN:
            pass
        elif position['size'] == 'NO SIZE':
            ws[f"{xl_size_pos[sizes[1]]}{i}"] = position['number']
        else:
            ws[f"{xl_size_pos[position['size']]}{i}"] = position['number']
        ws[f"N{i}"] = f"=SUM(C{i}:M{i})*B{i}"
    l = len(result['positions'])
    ws[f"N{l + 5}"] = f"=SUM(N{1}:N{l + 4})"
    ws[f"M{l + 5}"] = f"Итого: "

    save_workbook(wb, result['second_name'])


def fill_xl_group(result: dict):
    wb = load_workbook(result['group'])
    ws = copy_sheet_from_template(wb, 'Шаблон', "Final")
    wb.move_sheet("Final", -(len(wb.sheetnames) - 1))
    for i, position in enumerate(result['positions'], start=4):
        ws[f"A{i}"] = position['title']
        ws[f"B{i}"] = position['price']

        for letter_index, _ in enumerate(size_letters, start=3):
            ws.cell(column=letter_index, row=i).value = 0
            for sheet in wb.sheetnames:
                if sheet not in SHEET_EXCEPTIONS:
                    sheet_value = wb[sheet].cell(column=letter_index, row=i).value
                    if sheet_value is None:
                        pass
                    else:

                        ws.cell(column=letter_index, row=i).value += sheet_value
        for letter_index, _ in enumerate(size_letters, start=3):
            if ws.cell(column=letter_index, row=i).value == 0:
                ws.cell(column=letter_index, row=i).value = None
        ws[f"N{i}"] = f"=SUM(C{i}:M{i})*B{i}"

    l = len(result['positions'])
    ws[f"N{l + 5}"] = f"=SUM(N{1}:N{l + 4})"
    ws[f"M{l + 5}"] = f"Итого: "

    save_workbook(wb, result['group'])
